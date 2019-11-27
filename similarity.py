from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfVectorizer
from scipy.linalg import norm
import numpy as np
import distance
import gensim
import jieba
import xlrd
import openpyxl
import re

def add_space(s):
    return ' '.join(list(s))

def edit_distance(s1, s2):
    return distance.levenshtein(s1, s2)

def sparse_matrix_wordfreq(s1, s2):
    # 将字中间加入空格
    s1, s2 = add_space(s1), add_space(s2)
    # 转化为TF矩阵
    cv = CountVectorizer(tokenizer=lambda s: s.split())
    corpus = [s1, s2]
    vectors = cv.fit_transform(corpus).toarray()
    return vectors

def sparse_matrix_wordfreq_tfidf(s1, s2):
    # 将字中间加入空格
    s1, s2 = add_space(s1), add_space(s2)
    # 转化为TF矩阵
    cv = TfidfVectorizer(tokenizer=lambda s: s.split())
    corpus = [s1, s2]
    vectors = cv.fit_transform(corpus).toarray()
    return vectors

def jaccard_similarity(s1, s2):
    vectors = sparse_matrix_wordfreq(s1, s2)
    # 求交集
    numerator = np.sum(np.min(vectors, axis=0))
    # 求并集
    denominator = np.sum(np.max(vectors, axis=0))
    # 计算杰卡德系数
    return 1.0 * numerator / denominator

def cos_similarity(s1, s2):
    vectors = sparse_matrix_wordfreq(s1, s2)
    return np.dot(vectors[0], vectors[1]) / (norm(vectors[0]) * norm(vectors[1]))

def tfidf_similarity(s1, s2):
    vectors = sparse_matrix_wordfreq_tfidf(s1, s2)
    return np.dot(vectors[0], vectors[1]) / (norm(vectors[0]) * norm(vectors[1]))
 
def sentence_vector(s):
    words = jieba.lcut(s)
    v = np.zeros(300)
    words_include=[]
    for word in words:
        if word in model:
            words_include.append(word)
            v += model[word]
    v /= len(words_include)
    return v

def english_divided_str(str):
    sim_str =  ''.join(re.findall(r'[A-Za-z0-9/ /,/%/-/-/.]', str))
    return(sim_str)

def chinese_divided_str(str):
    sim_str = re.sub("[A-Za-z0-9\!\%\[\]\,\。\-\ \，\.\/]", "", str)
    return(sim_str)

def vector_similarity(s1, s2):
    v1, v2 = sentence_vector(s1), sentence_vector(s2)
    return np.dot(v1, v2) / (norm(v1) * norm(v2))

corpus=[]
workbook1 = openpyxl.load_workbook("./final_test2.xlsx")
worksheet1 = workbook1.get_sheet_by_name("english")
for cell in list(worksheet1.columns)[0]:
    corpus.append(english_divided_str(cell.value))

model_file = './word2vec/GoogleNews-vectors-negative300.bin'
model = gensim.models.KeyedVectors.load_word2vec_format(model_file, binary=True)

output1=[]
output2=[]
output3=[]
for i in range(0, len(corpus)):
    if i+1 > len(corpus):
        break
    else:
        for j in range(i+1, len(corpus)):
            output1.append(corpus[i])
            output2.append(corpus[j])
            output3.append(vector_similarity(corpus[i], corpus[j]))
print(output1, output2, output3)
workbook2 = openpyxl.load_workbook("./final_test2.xlsx")
worksheet2 = workbook2.get_sheet_by_name("similarity")

for i in range(len(output1)):
    worksheet2.cell(i+1, 1, output1[i])
    worksheet2.cell(i+1, 2, output2[i])
    worksheet2.cell(i+1, 3, output3[i])

workbook2.save(filename="./final_test2.xlsx")
