import xlrd
import openpyxl
import re

class extraction:

    def __init__(self, language, input_path, sheetname_in, 
                    column_in, column_out, strnum, output_path,
                        sheetname_out):
        self.column_in = column_in
        self.column_out = column_out
        self.strnum = strnum
        self.res = []
        self.new_res = []
        self.language = language
        self.input_path = input_path
        self.output_path = output_path
        self.sheetname_in = sheetname_in
        self.sheetname_out = sheetname_out

    def get_list_str(self, str):
        len_str = len(str)
        list_str = []
        for i in range(0, len_str-self.strnum-1):
            for j in range(i+self.strnum, i+self.strnum+1):
                two_str = str[i:j]
                list_str.append(two_str)
        list_str = list(set(list_str))
        return(list_str)

    def chinese_divided_str(self, str):
        sim_str = re.sub("[A-Za-z0-9\!\%\[\]\,\。\-\ \，\.\/]", "", str)
        return(sim_str)

    def english_divided_str(self, str):
        sim_str =  ''.join(re.findall(r'[A-Za-z0-9/ /,/%/-/-/.]', str))
        return(sim_str)

    def extract_str(self):
        workbook = openpyxl.load_workbook("./{0}.xlsx".format(self.input_path))
        worksheet = workbook.get_sheet_by_name(self.sheetname_in)
        for cell1 in list(worksheet.columns)[self.column_in]:
            for cell2 in list(worksheet.columns)[self.column_in]:
                if cell1.value == cell2.value:
                    continue
                if self.language == "chinese":
                    for x in self.get_list_str(self.chinese_divided_str(cell1.value)):
                        if x in self.get_list_str(self.chinese_divided_str(cell2.value)):
                            self.res.append(cell2.value)
                else:
                    for x in self.get_list_str(self.english_divided_str(cell1.value)):
                        if x in self.get_list_str(self.english_divided_str(cell2.value)):
                            self.res.append(cell2.value)
        self.new_res=list(set(self.res))
        self.new_res.sort(key=self.res.index)

    def save_sheet(self):
        workbook1 = openpyxl.load_workbook("./{0}.xlsx".format(self.output_path))
        worksheet1 = workbook1.get_sheet_by_name(self.sheetname_out)
        for i in range(len(self.new_res)):
            worksheet1.cell(i+1, self.column_out, self.new_res[i])
        workbook1.save(filename="./{0}.xlsx".format(self.output_path))

# e1 = extraction("chinese", "jira2", "raw_data", 2, "final_test1", "chinese")
# e1.extract_str()
# e2 = extraction("english", "final_test1", "english", 2, 1, 10, "final_test2", "english")
# e2.extract_str()
# e2.save_sheet()
# print(extraction.get_list_str(extraction.english_divided_str("PQ complained about heating terminal cover appearance not acceptable.")))